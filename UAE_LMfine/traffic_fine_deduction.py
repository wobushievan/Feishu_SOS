"""
Traffic fine monthly salary deduction processor.

Dependencies:
    pip install pandas openpyxl

How to run:
    python traffic_fine_deduction.py
    or double-click:
        run_traffic_fine_deduction.command (macOS)
        run_traffic_fine_deduction.bat (Windows)

Configuration:
    Update the constants in the "User configuration" section below.

Input file examples:
1. Historical fine workbook (first sheet will be read)
    Columns:
        Invoice | Ticket No | Fine date | Balance Due | ID | User | hub
    Optional historical helper columns that can already exist:
        Original Fine Amount
        Deducted Amount Cumulative
        Deducted Amount This Run
        Remaining Amount
        Deduction Month This Run

2. Driver-hub mapping workbook (first sheet will be read)
    Columns:
        ID | hub

3. Hub salary workbook (first sheet will be read)
    Columns:
        hub | Gross Salary | Basic Salary

Key logic:
    - The script rolls forward a historical fine table instead of recalculating from zero.
    - Driver processing is grouped by ID only.
    - Actual hub for deduction is always taken from the driver-hub mapping table.
    - Monthly deduction limit per driver = Gross Salary - Basic Salary of the mapped hub.
    - Deduction is allocated across the driver's unpaid fines in chronological order:
      Fine date, Invoice, Ticket No.
    - Driver-level configuration issues skip the whole driver for this run.
    - Row-level data issues skip only the affected fine row.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, Iterable, List, Set, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter


# =========================
# User configuration
# =========================
FINE_FILE_PATH = "司机罚款总表.xlsx"
HUB_SALARY_FILE_PATH = "网点工资映射.xlsx"
DRIVER_HUB_MAPPING_FILE_PATH = "司机网点映射.xlsx"
DEDUCTION_MONTH = "2026-03"
OUTPUT_FILE_PATH = f"traffic_fine_deduction_{DEDUCTION_MONTH.replace('-', '_')}.xlsx"




# =========================
# Constants
# =========================
SCRIPT_DIR = Path(__file__).resolve().parent
DETAIL_SHEET_NAME = "details"
SUMMARY_SHEET_NAME = "summary"
ERROR_SHEET_NAME = "errors"

DETAIL_REQUIRED_COLUMNS = [
    "Invoice",
    "Ticket No",
    "Fine date",
    "Balance Due",
    "ID",
    "User",
    "hub",
]
MAPPING_REQUIRED_COLUMNS = ["ID", "hub"]
SALARY_REQUIRED_COLUMNS = ["hub", "Gross Salary", "Basic Salary"]

DETAIL_ADDITIONAL_COLUMNS = [
    "Mapped Hub",
    "Original Fine Amount",
    "Deducted Amount Cumulative",
    "Deducted Amount This Run",
    "Remaining Amount",
    "Deduction Month This Run",
]

ERROR_COLUMNS = [
    "ID",
    "User",
    "Mapped Hub",
    "Source Hub",
    "Invoice",
    "Ticket No",
    "Error Type",
    "Error Detail",
]

SUMMARY_COLUMNS = [
    "ID",
    "User",
    "Mapped Hub",
    "Source Hub Values",
    "Opening Remaining Amount",
    "Monthly Deduction Limit",
    "Actual Deducted This Run",
    "Closing Remaining Amount",
    "Processing Status",
]

MONEY_COLUMNS_DETAILS = [
    "Balance Due",
    "Original Fine Amount",
    "Deducted Amount Cumulative",
    "Deducted Amount This Run",
    "Remaining Amount",
]
MONEY_COLUMNS_SUMMARY = [
    "Opening Remaining Amount",
    "Monthly Deduction Limit",
    "Actual Deducted This Run",
    "Closing Remaining Amount",
]
MONEY_COLUMNS_ERRORS: List[str] = []

TEMP_BALANCE_NUMERIC = "_Balance Due Numeric"
TEMP_FINE_DATE = "_Fine date Parsed"
TEMP_OPENING_REMAINING = "_Opening Remaining Amount"


class ProcessingError(Exception):
    """Raised when the workbook cannot be processed safely."""


def log(level: str, message: str) -> None:
    print(f"[{level}] {message}")


def log_info(message: str) -> None:
    log("INFO", message)


def log_warning(message: str) -> None:
    log("WARN", message)


def log_error(message: str) -> None:
    log("ERROR", message)


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_text_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    normalized.columns = [clean_text(column) for column in normalized.columns]
    for column in normalized.columns:
        normalized[column] = normalized[column].map(clean_text)
    return normalized


def parse_money_series(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("").map(clean_text).str.replace(",", "", regex=False)
    cleaned = cleaned.replace("", pd.NA)
    return pd.to_numeric(cleaned, errors="coerce")


def parse_fine_date_value(value: object) -> pd.Timestamp:
    text = clean_text(value)
    if not text:
        return pd.NaT

    if isinstance(value, datetime):
        return pd.Timestamp(value)

    if re.fullmatch(r"\d{4}[/-]\d{1,2}[/-]\d{1,2}", text):
        return pd.to_datetime(text, errors="coerce", yearfirst=True)

    if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{4}", text):
        return pd.to_datetime(text, errors="coerce", dayfirst=True)

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(parsed):
        return parsed
    return pd.to_datetime(text, errors="coerce", dayfirst=False)


def parse_fine_date_series(series: pd.Series) -> pd.Series:
    return series.map(parse_fine_date_value)


def round_money(value: object) -> float:
    if pd.isna(value):
        return 0.00
    return round(float(value) + 1e-9, 2)


def validate_deduction_month(deduction_month: str) -> None:
    if not re.fullmatch(r"\d{4}-\d{2}", deduction_month):
        raise ProcessingError(
            f"DEDUCTION_MONTH must use YYYY-MM format, got: {deduction_month}"
        )


def resolve_path(path_value: str) -> Path:
    path = Path(path_value).expanduser()
    if path.is_absolute():
        return path
    return SCRIPT_DIR / path


def build_error_record(
    *,
    driver_id: str = "",
    user: str = "",
    mapped_hub: str = "",
    source_hub: str = "",
    invoice: str = "",
    ticket_no: str = "",
    error_type: str,
    error_detail: str,
) -> Dict[str, str]:
    return {
        "ID": clean_text(driver_id),
        "User": clean_text(user),
        "Mapped Hub": clean_text(mapped_hub),
        "Source Hub": clean_text(source_hub),
        "Invoice": clean_text(invoice),
        "Ticket No": clean_text(ticket_no),
        "Error Type": clean_text(error_type),
        "Error Detail": clean_text(error_detail),
    }


def append_error(
    errors: List[Dict[str, str]],
    *,
    driver_id: str = "",
    user: str = "",
    mapped_hub: str = "",
    source_hub: str = "",
    invoice: str = "",
    ticket_no: str = "",
    error_type: str,
    error_detail: str,
    severity: str = "ERROR",
) -> None:
    errors.append(
        build_error_record(
            driver_id=driver_id,
            user=user,
            mapped_hub=mapped_hub,
            source_hub=source_hub,
            invoice=invoice,
            ticket_no=ticket_no,
            error_type=error_type,
            error_detail=error_detail,
        )
    )
    if severity.upper() == "WARN":
        log_warning(error_detail)
    else:
        log_error(error_detail)


def load_input_files() -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    log_info("Loading input workbooks from the configured file paths.")
    fine_path = resolve_path(FINE_FILE_PATH)
    hub_salary_path = resolve_path(HUB_SALARY_FILE_PATH)
    driver_hub_path = resolve_path(DRIVER_HUB_MAPPING_FILE_PATH)

    try:
        fine_df = pd.read_excel(fine_path, sheet_name=0, dtype=str)
        hub_salary_df = pd.read_excel(hub_salary_path, sheet_name=0, dtype=str)
        driver_hub_df = pd.read_excel(driver_hub_path, sheet_name=0, dtype=str)
    except FileNotFoundError as exc:
        raise ProcessingError(f"Input file not found: {exc.filename}") from exc
    except Exception as exc:  # pragma: no cover - defensive path
        raise ProcessingError(f"Failed to read input workbooks: {exc}") from exc

    return (
        normalize_text_dataframe(fine_df),
        normalize_text_dataframe(hub_salary_df),
        normalize_text_dataframe(driver_hub_df),
    )


def validate_required_columns(
    df: pd.DataFrame, required_columns: Iterable[str], sheet_name: str
) -> None:
    missing_columns = [column for column in required_columns if column not in df.columns]
    if missing_columns:
        missing_str = ", ".join(missing_columns)
        raise ProcessingError(
            f"Missing required columns in '{sheet_name}' sheet: {missing_str}"
        )


def validate_driver_hub_mapping(
    mapping_df: pd.DataFrame, errors: List[Dict[str, str]]
) -> Dict[str, str]:
    log_info("Validating driver-hub mapping table.")
    invalid_driver_status: Dict[str, str] = {}

    for row in mapping_df.itertuples(index=False):
        driver_id = clean_text(getattr(row, "ID"))
        hub = clean_text(getattr(row, "hub"))

        if not driver_id:
            append_error(
                errors,
                error_type="Invalid mapping row",
                error_detail="Driver-hub mapping row has empty ID and will be ignored.",
            )
            continue

        if not hub:
            invalid_driver_status[driver_id] = "Skipped due to missing driver-hub mapping"
            append_error(
                errors,
                driver_id=driver_id,
                error_type="Driver hub missing in mapping",
                error_detail=(
                    f"Driver ID '{driver_id}' has an empty hub in the mapping table. "
                    "The driver will be skipped for this run."
                ),
            )

    valid_mapping_rows = mapping_df[
        mapping_df["ID"].ne("") & mapping_df["hub"].ne("")
    ].copy()
    distinct_hubs = (
        valid_mapping_rows.groupby("ID")["hub"].nunique(dropna=True).reset_index()
    )
    multi_hub_drivers = distinct_hubs.loc[distinct_hubs["hub"] > 1, "ID"].tolist()

    for driver_id in multi_hub_drivers:
        invalid_driver_status[driver_id] = "Skipped due to multiple hubs in mapping"
        driver_rows = valid_mapping_rows.loc[valid_mapping_rows["ID"] == driver_id, "hub"]
        hub_values = ", ".join(sorted(driver_rows.unique()))
        append_error(
            errors,
            driver_id=driver_id,
            error_type="Driver mapped to multiple hubs",
            error_detail=(
                f"Driver ID '{driver_id}' is mapped to multiple hubs in the mapping table: "
                f"{hub_values}. The driver will be skipped for this run."
            ),
        )

    return invalid_driver_status


def build_driver_hub_mapping(
    mapping_df: pd.DataFrame, invalid_driver_status: Dict[str, str]
) -> Dict[str, str]:
    log_info("Building driver -> mapped hub lookup.")
    valid_mapping_df = mapping_df[
        mapping_df["ID"].ne("")
        & mapping_df["hub"].ne("")
        & ~mapping_df["ID"].isin(invalid_driver_status.keys())
    ].copy()
    valid_mapping_df = valid_mapping_df.drop_duplicates(subset=["ID"], keep="first")
    return dict(zip(valid_mapping_df["ID"], valid_mapping_df["hub"]))


def build_hub_salary_limit_mapping(
    salary_df: pd.DataFrame, errors: List[Dict[str, str]]
) -> Tuple[Dict[str, float], Dict[str, str]]:
    log_info("Building hub -> monthly deduction limit lookup.")
    salary_limit_map: Dict[str, float] = {}
    invalid_hub_status: Dict[str, str] = {}

    salary_df = salary_df.copy()
    salary_df["_gross_numeric"] = parse_money_series(salary_df["Gross Salary"])
    salary_df["_basic_numeric"] = parse_money_series(salary_df["Basic Salary"])

    duplicate_hubs = (
        salary_df.loc[salary_df["hub"].ne(""), "hub"]
        .value_counts()
        .loc[lambda series: series > 1]
        .index.tolist()
    )
    for hub in duplicate_hubs:
        invalid_hub_status[hub] = "Skipped due to duplicate hub salary config"
        append_error(
            errors,
            mapped_hub=hub,
            error_type="Duplicate hub salary config",
            error_detail=(
                f"Hub '{hub}' appears multiple times in the salary config. "
                "Drivers mapped to this hub will be skipped for this run."
            ),
        )

    for _, row in salary_df.iterrows():
        hub = clean_text(row["hub"])
        gross_salary = row["_gross_numeric"]
        basic_salary = row["_basic_numeric"]

        if not hub:
            append_error(
                errors,
                error_type="Invalid salary row",
                error_detail="Hub salary config row has empty hub and will be ignored.",
            )
            continue

        if hub in invalid_hub_status:
            continue

        if pd.isna(gross_salary) or pd.isna(basic_salary):
            invalid_hub_status[hub] = "Skipped due to invalid salary config"
            append_error(
                errors,
                mapped_hub=hub,
                error_type="Invalid salary config",
                error_detail=(
                    f"Hub '{hub}' has non-numeric Gross Salary or Basic Salary. "
                    "Drivers mapped to this hub will be skipped for this run."
                ),
            )
            continue

        gross_salary = round_money(gross_salary)
        basic_salary = round_money(basic_salary)
        if gross_salary < basic_salary:
            invalid_hub_status[hub] = "Skipped due to invalid salary config"
            append_error(
                errors,
                mapped_hub=hub,
                error_type="Invalid salary config",
                error_detail=(
                    f"Hub '{hub}' has Gross Salary ({gross_salary:.2f}) lower than "
                    f"Basic Salary ({basic_salary:.2f}). Drivers mapped to this hub "
                    "will be skipped for this run."
                ),
            )
            continue

        salary_limit_map[hub] = round_money(gross_salary - basic_salary)

    return salary_limit_map, invalid_hub_status


def initialize_or_normalize_columns(
    fine_df: pd.DataFrame, errors: List[Dict[str, str]]
) -> Tuple[pd.DataFrame, Set[int]]:
    log_info("Initializing and normalizing deduction helper columns.")
    normalized_df = fine_df.copy()
    invalid_row_indices: Set[int] = set()

    for column in DETAIL_ADDITIONAL_COLUMNS:
        if column not in normalized_df.columns:
            normalized_df[column] = ""

    normalized_df[TEMP_BALANCE_NUMERIC] = parse_money_series(normalized_df["Balance Due"])

    original_numeric = parse_money_series(normalized_df["Original Fine Amount"])
    cumulative_numeric = parse_money_series(normalized_df["Deducted Amount Cumulative"])
    remaining_numeric = parse_money_series(normalized_df["Remaining Amount"])

    new_ticket_mask = (
        normalized_df["Original Fine Amount"].eq("")
        | normalized_df["Deducted Amount Cumulative"].eq("")
        | normalized_df["Remaining Amount"].eq("")
    )

    helper_numeric_invalid_mask = (
        ~new_ticket_mask
        & (
            original_numeric.isna()
            | cumulative_numeric.isna()
            | remaining_numeric.isna()
        )
    )

    for idx in normalized_df.index[helper_numeric_invalid_mask]:
        invalid_row_indices.add(int(idx))
        append_error(
            errors,
            driver_id=normalized_df.at[idx, "ID"],
            user=normalized_df.at[idx, "User"],
            source_hub=normalized_df.at[idx, "hub"],
            invoice=normalized_df.at[idx, "Invoice"],
            ticket_no=normalized_df.at[idx, "Ticket No"],
            error_type="Invalid historical deduction columns",
            error_detail=(
                f"Row with Invoice '{normalized_df.at[idx, 'Invoice']}' and Ticket No "
                f"'{normalized_df.at[idx, 'Ticket No']}' has non-numeric historical "
                "deduction helper values and will be skipped."
            ),
        )

    normalized_df["Original Fine Amount"] = original_numeric
    normalized_df["Deducted Amount Cumulative"] = cumulative_numeric
    normalized_df["Remaining Amount"] = remaining_numeric

    normalized_df.loc[new_ticket_mask, "Original Fine Amount"] = normalized_df.loc[
        new_ticket_mask, TEMP_BALANCE_NUMERIC
    ]
    normalized_df.loc[new_ticket_mask, "Deducted Amount Cumulative"] = 0.00
    normalized_df.loc[new_ticket_mask, "Remaining Amount"] = normalized_df.loc[
        new_ticket_mask, TEMP_BALANCE_NUMERIC
    ]

    normalized_df["Deducted Amount This Run"] = 0.00
    normalized_df["Deduction Month This Run"] = ""
    normalized_df["Mapped Hub"] = normalized_df["Mapped Hub"].fillna("").map(clean_text)

    formula_check_mask = (
        ~new_ticket_mask
        & ~helper_numeric_invalid_mask
        & (
            (
                normalized_df["Original Fine Amount"]
                - normalized_df["Deducted Amount Cumulative"]
                - normalized_df["Remaining Amount"]
            )
            .round(2)
            .abs()
            > 0.01
        )
    )
    for idx in normalized_df.index[formula_check_mask]:
        invalid_row_indices.add(int(idx))
        append_error(
            errors,
            driver_id=normalized_df.at[idx, "ID"],
            user=normalized_df.at[idx, "User"],
            source_hub=normalized_df.at[idx, "hub"],
            invoice=normalized_df.at[idx, "Invoice"],
            ticket_no=normalized_df.at[idx, "Ticket No"],
            error_type="Historical deduction mismatch",
            error_detail=(
                f"Row with Invoice '{normalized_df.at[idx, 'Invoice']}' and Ticket No "
                f"'{normalized_df.at[idx, 'Ticket No']}' does not satisfy "
                "'Original Fine Amount = Deducted Amount Cumulative + Remaining Amount' "
                "and will be skipped."
            ),
        )

    money_columns = [
        "Original Fine Amount",
        "Deducted Amount Cumulative",
        "Deducted Amount This Run",
        "Remaining Amount",
    ]
    for column in money_columns:
        normalized_df[column] = normalized_df[column].map(round_money)

    normalized_df[TEMP_OPENING_REMAINING] = normalized_df["Remaining Amount"].map(round_money)
    return normalized_df, invalid_row_indices


def validate_fine_sheet_data(
    fine_df: pd.DataFrame,
    driver_hub_map: Dict[str, str],
    errors: List[Dict[str, str]],
    invalid_row_indices: Set[int],
) -> Tuple[pd.DataFrame, Set[int]]:
    log_info("Validating fine detail data.")
    validated_df = fine_df.copy()

    validated_df[TEMP_FINE_DATE] = parse_fine_date_series(validated_df["Fine date"])

    missing_id_mask = validated_df["ID"].eq("")
    for idx in validated_df.index[missing_id_mask]:
        invalid_row_indices.add(int(idx))
        append_error(
            errors,
            user=validated_df.at[idx, "User"],
            source_hub=validated_df.at[idx, "hub"],
            invoice=validated_df.at[idx, "Invoice"],
            ticket_no=validated_df.at[idx, "Ticket No"],
            error_type="Missing driver ID",
            error_detail=(
                f"Row with Invoice '{validated_df.at[idx, 'Invoice']}' and Ticket No "
                f"'{validated_df.at[idx, 'Ticket No']}' has an empty driver ID and "
                "will be skipped."
            ),
        )

    invalid_balance_mask = validated_df[TEMP_BALANCE_NUMERIC].isna()
    for idx in validated_df.index[invalid_balance_mask]:
        invalid_row_indices.add(int(idx))
        append_error(
            errors,
            driver_id=validated_df.at[idx, "ID"],
            user=validated_df.at[idx, "User"],
            source_hub=validated_df.at[idx, "hub"],
            invoice=validated_df.at[idx, "Invoice"],
            ticket_no=validated_df.at[idx, "Ticket No"],
            error_type="Invalid Balance Due",
            error_detail=(
                f"Row with Invoice '{validated_df.at[idx, 'Invoice']}' and Ticket No "
                f"'{validated_df.at[idx, 'Ticket No']}' has non-numeric Balance Due "
                "and will be skipped."
            ),
        )

    invalid_date_mask = validated_df[TEMP_FINE_DATE].isna()
    for idx in validated_df.index[invalid_date_mask]:
        invalid_row_indices.add(int(idx))
        append_error(
            errors,
            driver_id=validated_df.at[idx, "ID"],
            user=validated_df.at[idx, "User"],
            source_hub=validated_df.at[idx, "hub"],
            invoice=validated_df.at[idx, "Invoice"],
            ticket_no=validated_df.at[idx, "Ticket No"],
            error_type="Invalid Fine date",
            error_detail=(
                f"Row with Invoice '{validated_df.at[idx, 'Invoice']}' and Ticket No "
                f"'{validated_df.at[idx, 'Ticket No']}' has an unparseable Fine date "
                "and will be skipped."
            ),
        )

    missing_unique_key_mask = validated_df["Invoice"].eq("") | validated_df["Ticket No"].eq("")
    for idx in validated_df.index[missing_unique_key_mask]:
        invalid_row_indices.add(int(idx))
        append_error(
            errors,
            driver_id=validated_df.at[idx, "ID"],
            user=validated_df.at[idx, "User"],
            source_hub=validated_df.at[idx, "hub"],
            invoice=validated_df.at[idx, "Invoice"],
            ticket_no=validated_df.at[idx, "Ticket No"],
            error_type="Missing Invoice or Ticket No",
            error_detail=(
                f"Row for driver '{validated_df.at[idx, 'ID']}' is missing Invoice or "
                "Ticket No and will be skipped."
            ),
        )

    validated_df["_unique_key"] = (
        validated_df["Invoice"].fillna("") + "||" + validated_df["Ticket No"].fillna("")
    )
    duplicate_mask = validated_df["_unique_key"].duplicated(keep=False) & ~missing_unique_key_mask
    duplicate_keys = validated_df.loc[duplicate_mask, "_unique_key"].unique().tolist()
    for key in duplicate_keys:
        duplicate_rows = validated_df.loc[validated_df["_unique_key"] == key]
        invoice_value = duplicate_rows["Invoice"].iloc[0]
        ticket_value = duplicate_rows["Ticket No"].iloc[0]
        for idx in duplicate_rows.index:
            invalid_row_indices.add(int(idx))
            append_error(
                errors,
                driver_id=validated_df.at[idx, "ID"],
                user=validated_df.at[idx, "User"],
                source_hub=validated_df.at[idx, "hub"],
                invoice=validated_df.at[idx, "Invoice"],
                ticket_no=validated_df.at[idx, "Ticket No"],
                error_type="Duplicate Invoice + Ticket No",
                error_detail=(
                    f"Duplicate fine key detected for Invoice '{invoice_value}' and "
                    f"Ticket No '{ticket_value}'. All rows with this key will be skipped."
                ),
            )

    validated_df["Mapped Hub"] = validated_df["ID"].map(driver_hub_map).fillna("")

    hub_mismatch_mask = (
        validated_df["Mapped Hub"].ne("")
        & validated_df["hub"].ne("")
        & validated_df["Mapped Hub"].ne(validated_df["hub"])
    )
    for idx in validated_df.index[hub_mismatch_mask]:
        append_error(
            errors,
            driver_id=validated_df.at[idx, "ID"],
            user=validated_df.at[idx, "User"],
            mapped_hub=validated_df.at[idx, "Mapped Hub"],
            source_hub=validated_df.at[idx, "hub"],
            invoice=validated_df.at[idx, "Invoice"],
            ticket_no=validated_df.at[idx, "Ticket No"],
            error_type="Hub mismatch warning",
            error_detail=(
                f"Source hub '{validated_df.at[idx, 'hub']}' does not match mapped hub "
                f"'{validated_df.at[idx, 'Mapped Hub']}' for driver "
                f"'{validated_df.at[idx, 'ID']}'. Processing will use the mapped hub."
            ),
            severity="WARN",
        )

    validated_df = validated_df.drop(columns=["_unique_key"])
    return validated_df, invalid_row_indices


def process_driver_deductions(
    fine_df: pd.DataFrame,
    driver_hub_map: Dict[str, str],
    invalid_driver_status: Dict[str, str],
    hub_salary_limit_map: Dict[str, float],
    invalid_hub_status: Dict[str, str],
    errors: List[Dict[str, str]],
    invalid_row_indices: Set[int],
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    log_info("Processing monthly deductions by driver.")
    processed_df = fine_df.copy()
    driver_status_map: Dict[str, str] = {}

    driver_ids = sorted(driver_id for driver_id in processed_df["ID"].unique() if driver_id)

    for driver_id in driver_ids:
        driver_rows = processed_df.loc[processed_df["ID"] == driver_id]
        user = next((value for value in driver_rows["User"] if value), "")
        source_hub_values = sorted(value for value in driver_rows["hub"].unique() if value)
        source_hub_text = " | ".join(source_hub_values)

        if driver_id in invalid_driver_status:
            driver_status_map[driver_id] = invalid_driver_status[driver_id]
            continue

        mapped_hub = driver_hub_map.get(driver_id, "")
        if not mapped_hub:
            driver_status_map[driver_id] = "Skipped due to missing driver-hub mapping"
            append_error(
                errors,
                driver_id=driver_id,
                user=user,
                source_hub=source_hub_text,
                error_type="Driver missing in mapping",
                error_detail=(
                    f"Driver ID '{driver_id}' was not found in the driver-hub mapping "
                    "table and will be skipped for this run."
                ),
            )
            continue

        processed_df.loc[driver_rows.index, "Mapped Hub"] = mapped_hub

        if mapped_hub in invalid_hub_status:
            driver_status_map[driver_id] = invalid_hub_status[mapped_hub]
            append_error(
                errors,
                driver_id=driver_id,
                user=user,
                mapped_hub=mapped_hub,
                source_hub=source_hub_text,
                error_type="Driver skipped due to hub salary config",
                error_detail=(
                    f"Driver '{driver_id}' mapped to hub '{mapped_hub}' was skipped: "
                    f"{invalid_hub_status[mapped_hub]}."
                ),
            )
            continue

        if mapped_hub not in hub_salary_limit_map:
            driver_status_map[driver_id] = "Skipped due to missing hub salary config"
            append_error(
                errors,
                driver_id=driver_id,
                user=user,
                mapped_hub=mapped_hub,
                source_hub=source_hub_text,
                error_type="Missing hub salary config",
                error_detail=(
                    f"Mapped hub '{mapped_hub}' for driver '{driver_id}' was not found "
                    "in the salary config table. The driver will be skipped for this run."
                ),
            )
            continue

        driver_status_map[driver_id] = "Processed"
        available_limit = round_money(hub_salary_limit_map[mapped_hub])
        valid_indices = [
            idx
            for idx in driver_rows.index
            if idx not in invalid_row_indices
            and round_money(processed_df.at[idx, "Remaining Amount"]) > 0
        ]
        if not valid_indices or available_limit <= 0:
            continue

        sortable_rows = processed_df.loc[valid_indices].sort_values(
            by=[TEMP_FINE_DATE, "Invoice", "Ticket No"],
            ascending=[True, True, True],
        )

        for idx in sortable_rows.index:
            if available_limit <= 0:
                break
            remaining_amount = round_money(processed_df.at[idx, "Remaining Amount"])
            if remaining_amount <= 0:
                continue

            deduct_amount = round_money(min(available_limit, remaining_amount))
            if deduct_amount <= 0:
                continue

            cumulative_amount = round_money(
                processed_df.at[idx, "Deducted Amount Cumulative"] + deduct_amount
            )
            closing_remaining = round_money(remaining_amount - deduct_amount)

            processed_df.at[idx, "Deducted Amount This Run"] = deduct_amount
            processed_df.at[idx, "Deducted Amount Cumulative"] = cumulative_amount
            processed_df.at[idx, "Remaining Amount"] = closing_remaining
            processed_df.at[idx, "Deduction Month This Run"] = DEDUCTION_MONTH

            available_limit = round_money(available_limit - deduct_amount)

    return processed_df, driver_status_map


def build_summary_sheet(
    fine_df: pd.DataFrame,
    driver_hub_map: Dict[str, str],
    hub_salary_limit_map: Dict[str, float],
    driver_status_map: Dict[str, str],
) -> pd.DataFrame:
    log_info("Building summary sheet.")
    summary_rows: List[Dict[str, object]] = []

    driver_ids = sorted(driver_id for driver_id in fine_df["ID"].unique() if driver_id)
    for driver_id in driver_ids:
        driver_rows = fine_df.loc[fine_df["ID"] == driver_id]
        user = next((value for value in driver_rows["User"] if value), "")
        mapped_hub = driver_hub_map.get(driver_id, "")
        if not mapped_hub and "Mapped Hub" in driver_rows.columns:
            mapped_hub = next((value for value in driver_rows["Mapped Hub"] if value), "")

        source_hub_values = sorted(value for value in driver_rows["hub"].unique() if value)
        source_hub_text = " | ".join(source_hub_values)

        opening_remaining = round_money(driver_rows[TEMP_OPENING_REMAINING].sum())
        actual_deducted = round_money(driver_rows["Deducted Amount This Run"].sum())
        closing_remaining = round_money(driver_rows["Remaining Amount"].sum())
        status = driver_status_map.get(driver_id, "Processed")
        monthly_limit = ""
        if mapped_hub and mapped_hub in hub_salary_limit_map and status == "Processed":
            monthly_limit = round_money(hub_salary_limit_map[mapped_hub])

        summary_rows.append(
            {
                "ID": driver_id,
                "User": user,
                "Mapped Hub": mapped_hub,
                "Source Hub Values": source_hub_text,
                "Opening Remaining Amount": opening_remaining,
                "Monthly Deduction Limit": monthly_limit,
                "Actual Deducted This Run": actual_deducted,
                "Closing Remaining Amount": closing_remaining,
                "Processing Status": status,
            }
        )

    summary_df = pd.DataFrame(summary_rows, columns=SUMMARY_COLUMNS)
    if summary_df.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    for column in [
        "Opening Remaining Amount",
        "Actual Deducted This Run",
        "Closing Remaining Amount",
    ]:
        summary_df[column] = summary_df[column].map(round_money)

    summary_df = summary_df.sort_values(by=["ID"]).reset_index(drop=True)
    return summary_df


def build_errors_sheet(errors: List[Dict[str, str]]) -> pd.DataFrame:
    log_info("Building errors sheet.")
    if not errors:
        return pd.DataFrame(columns=ERROR_COLUMNS)
    return pd.DataFrame(errors, columns=ERROR_COLUMNS)


def format_worksheet(worksheet, money_columns: List[str]) -> None:
    header_map = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_name in money_columns:
        if column_name not in header_map:
            continue
        column_letter = get_column_letter(header_map[column_name])
        for cell in worksheet[column_letter][1:]:
            cell.number_format = "#,##0.00"

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def save_output_workbook(
    details_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    errors_df: pd.DataFrame,
    output_file_path: str,
) -> None:
    output_path = resolve_path(output_file_path)
    log_info(f"Saving output workbook to: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        details_df.to_excel(writer, sheet_name=DETAIL_SHEET_NAME, index=False)
        summary_df.to_excel(writer, sheet_name=SUMMARY_SHEET_NAME, index=False)
        errors_df.to_excel(writer, sheet_name=ERROR_SHEET_NAME, index=False)

        format_worksheet(writer.sheets[DETAIL_SHEET_NAME], MONEY_COLUMNS_DETAILS)
        format_worksheet(writer.sheets[SUMMARY_SHEET_NAME], MONEY_COLUMNS_SUMMARY)
        format_worksheet(writer.sheets[ERROR_SHEET_NAME], MONEY_COLUMNS_ERRORS)


def main() -> None:
    validate_deduction_month(DEDUCTION_MONTH)

    errors: List[Dict[str, str]] = []

    fine_df, hub_salary_df, driver_hub_df = load_input_files()
    validate_required_columns(fine_df, DETAIL_REQUIRED_COLUMNS, "historical fine")
    validate_required_columns(hub_salary_df, SALARY_REQUIRED_COLUMNS, "hub salary")
    validate_required_columns(driver_hub_df, MAPPING_REQUIRED_COLUMNS, "driver-hub mapping")

    invalid_driver_status = validate_driver_hub_mapping(driver_hub_df, errors)
    driver_hub_map = build_driver_hub_mapping(driver_hub_df, invalid_driver_status)
    hub_salary_limit_map, invalid_hub_status = build_hub_salary_limit_mapping(
        hub_salary_df, errors
    )

    fine_df, invalid_row_indices = initialize_or_normalize_columns(fine_df, errors)
    fine_df, invalid_row_indices = validate_fine_sheet_data(
        fine_df, driver_hub_map, errors, invalid_row_indices
    )
    fine_df, driver_status_map = process_driver_deductions(
        fine_df=fine_df,
        driver_hub_map=driver_hub_map,
        invalid_driver_status=invalid_driver_status,
        hub_salary_limit_map=hub_salary_limit_map,
        invalid_hub_status=invalid_hub_status,
        errors=errors,
        invalid_row_indices=invalid_row_indices,
    )

    for column in [
        "Original Fine Amount",
        "Deducted Amount Cumulative",
        "Deducted Amount This Run",
        "Remaining Amount",
    ]:
        fine_df[column] = fine_df[column].map(round_money)

    summary_df = build_summary_sheet(
        fine_df=fine_df,
        driver_hub_map=driver_hub_map,
        hub_salary_limit_map=hub_salary_limit_map,
        driver_status_map=driver_status_map,
    )
    errors_df = build_errors_sheet(errors)

    details_output_df = fine_df.drop(
        columns=[TEMP_BALANCE_NUMERIC, TEMP_FINE_DATE, TEMP_OPENING_REMAINING],
        errors="ignore",
    )

    ordered_columns = list(fine_df.columns)
    for temp_column in [TEMP_BALANCE_NUMERIC, TEMP_FINE_DATE, TEMP_OPENING_REMAINING]:
        if temp_column in ordered_columns:
            ordered_columns.remove(temp_column)
    details_output_df = details_output_df[ordered_columns]

    save_output_workbook(
        details_df=details_output_df,
        summary_df=summary_df,
        errors_df=errors_df,
        output_file_path=OUTPUT_FILE_PATH,
    )

    log_info(
        "Processing completed successfully. "
        f"Detail rows: {len(details_output_df)}, "
        f"summary rows: {len(summary_df)}, "
        f"error rows: {len(errors_df)}."
    )


if __name__ == "__main__":
    try:
        main()
    except ProcessingError as exc:
        log_error(str(exc))
        sys.exit(1)
    except Exception as exc:  # pragma: no cover - defensive path
        log_error(f"Unexpected failure: {exc}")
        raise
